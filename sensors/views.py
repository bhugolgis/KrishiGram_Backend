
from io import BytesIO , StringIO
import json
from openpyxl import load_workbook
import pytz
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser
from .serializers import *
from rest_framework import status
from rest_framework.response import Response
from rest_framework import generics
from rest_framework.exceptions import NotFound
from django.contrib.gis.geos import Point
from django.http import HttpResponse
import pandas as pd
from django.contrib.gis.db.models.functions import AsGeoJSON
import csv
import datetime
# from datetime import datetime
from .models import *
from rest_framework.filters import SearchFilter
from dateutil import parser
from datetime import timedelta, timezone




# Create your views here.

# sensor property post api
class PostDevicesAPIview(generics.GenericAPIView):
    serializer_class= PostDevicesAPISerailzer
    parser_classes=[MultiPartParser]

    def post(self, request,format=None):

        serializer = self.get_serializer(data=request.data)
        lat=float(request.data['latitude'])
        long=float(request.data['longitude'])
        location=Point(long,lat,srid=4326)
        if serializer.is_valid():
            serializer.save(location = location)

            return Response( { 'status' : 'success',
                                'message':'Device Created succesfully'}
                                ,status=status.HTTP_201_CREATED)
        else:
            key, value =list(serializer.errors.items())[0]
            error_message = key+" ,"+value[0]
            return Response({'status': 'error',
                             'message': error_message} , status=400)





# device_id wise fetch data
class SensorDataView(APIView):
    def get(self,request,device_id,format=None):
        try:
            sensor_data = SoilSensor.objects.filter(devices__device_id=device_id)
            serializer = SensordataviewSerializer(sensor_data,many=True)
            return Response(serializer.data)
        except SoilSensor.DoesNotExist:
            return Response({
                'error': 'Device not found'
            }, status=404)





class SensorList(APIView):
    def get(self, request, sensor_type=None, start_date=None, end_date=None):
        # Retrieve the list of Sensorproperty objects
        sensor_properties = devices.objects.all()

        # Get a list of unique sensor types
        sensor_types = set([prop.sensor_type for prop in sensor_properties])


        if not sensor_type:
            # Return the list of sensor types
            return Response(sensor_types)

        # Filter the Sensorproperty objects by sensor_type
        sensor_properties = devices.objects.filter(sensor_type=sensor_type)
        # print(sensor_properties)

        # Filter the Sensor objects by property and created_at between start_date and end_date for each Sensorproperty
        sensors = []
        for sensor_property in sensor_properties:
            sensor_queryset = SoilSensor.objects.filter(devices=sensor_property,
                                                    Timestamp__date__range=(start_date, end_date))
            # print(sensor_queryset)
            sensors.extend(sensor_queryset)

        # Check if any Sensor objects were found
        if not sensors:
            return Response({"message": "No Sensor data found for the specified type and date."},
                            status=status.HTTP_404_NOT_FOUND)
        
        for sensor in sensors:
            formatted_timestamp = sensor.Timestamp.strftime('%Y-%m-%d %I:%M:%S %p')
            sensor.Timestamp = formatted_timestamp


        # Serialize the data and return a JSON response
        serializer = SensorlistSerializer(sensors, many=True)
        return Response(serializer.data)




# class GeosensorList(APIView):
#     def get(self, request, sensor_type=None, start_date=None, end_date=None):
#         # Retrieve the list of Sensorproperty objects
#         sensor_properties = devices.objects.all()

#         if not sensor_type:
#             # Return the list of sensor types
#             sensor_types = set([prop.sensor_type for prop in sensor_properties])
#             return Response(sensor_types)

#         # Filter the Sensorproperty objects by sensor_type
#         sensor_properties = devices.objects.filter(sensor_type=sensor_type)

#         # Filter the Sensor objects by property and created_at between start_date and end_date for each Sensorproperty
#         sensors = []
#         for sensor_property in sensor_properties:
#             sensor_queryset = SoilSensor.objects.filter(devices=sensor_property, Timestamp__date__range=(start_date, end_date))
#             sensors.extend(sensor_queryset)

#         # Check if any Sensor objects were found
#         if not sensors:
#             return Response({"message": "No Sensor data found for the specified type and date."}, status=status.HTTP_404_NOT_FOUND)

#         # Serialize the data in GeoJSON format
#         data = {
#             'type': 'FeatureCollection',
#             'features': []
#         }
#         for sensor in sensors:
#             point = sensor.devices.location
#             feature = {
#                 'type': 'Feature',
#                 'geometry': {
#                     'type': 'Point',
#                     'coordinates': [point.x, point.y]
#                 },
#                 'properties': SensorlistSerializer(sensor).data
#             }
#             data['features'].append(feature)

#         # Adjust the timestamp format
#         for feature in data['features']:
#             timestamp = feature['properties']['Timestamp']
#             formatted_timestamp = datetime.datetime.strptime(timestamp, '%Y-%m-%dT%H:%M:%S.%fZ')
#             localized_timestamp = formatted_timestamp.replace(tzinfo=datetime.timezone.utc).astimezone(pytz.timezone('UTC'))
#             feature['properties']['Timestamp'] = localized_timestamp.strftime('%I:%M:%S %p %d-%m-%Y')

#         # Return the GeoJSON response
#         return Response(data)





#  Advisory
class AdvisoryExcelInsertxlsx(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = uploadExcelserializer
    def post(self, request, format=None):
        # try:
        if 'excel_file' not in request.FILES:
            return Response({"status": "error",
                              "message": "No file uploaded."}, status=400)
        excel_file = request.FILES["excel_file"]
        Advisory_date = request.data['Advisory_date']
        if excel_file.size == 0 or excel_file.name.endswith(".xlsx") != True:

            return Response({"status": "error",
                             "message": "only .xlsx file is supported."},status=400)
        date = request.data["Advisory_date"]
        already_exist = Advisory.objects.filter(Advisory_date = date).exists()
        if already_exist:
            return Response ({"status": "error",
                              "message": f"Advisory Already Uploaded For Selected Date - {date}"} , status= 400)

        workbook = load_workbook(filename=excel_file)

        # print('------220',workbook)
        sheet_name = workbook.sheetnames[0]
        # print(sheet_name)
        worksheet = workbook[sheet_name]
        # print(worksheet)
        data_list = []

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                break
            data = Advisory(crop_name=row[0], Stage=row[1], Agromet_Advisory=row[2] , Advisory_date = Advisory_date )
            data_list.append(data)
        # print(data_list)
        Advisory.objects.bulk_create(data_list)
        return Response({"status": "Success","message": "Successfully Uploaded."})


class AdvisoryList(generics.GenericAPIView):
    queryset = Advisory.objects.all()
    serializer_class = AdvisorySerializer

    def get(self, request, advisory_date):
        advisories = self.get_queryset().filter(Advisory_date = advisory_date)
        if not advisories:
            return Response({
                "message" : "There is no Advisory data Found for selected Date",
                "status" : "error",
            } , status= 400)
            # last_created_date =  self.get_queryset().Advisory_date
            # data =  self.get_queryset().filter(Advisory_date=advisory_date)
            # serializer = self.get_serializer(data, many=True).data

            # return Response({"message":"Data Fetched Successfully",
            #                  "status" : "success",
            #                  "uploaded_at" : last_created_date.date(),
            #                  "data":serializer})
        
        serializer = self.get_serializer(advisories, many=True).data
        return Response({"message":"Data Fetched Successfully",
                        "status" : "success",
                        "advisory_date" : advisory_date,
                        "data":serializer})


class SensorUploadView(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = JsonuploadSerializer

    def post(self, request):
        serializer = JsonuploadSerializer(data=request.data)
        if serializer.is_valid():
            file = serializer.validated_data['file']
            json_data = file.read().decode('utf-8')

            sensor_data = json.loads(json_data)
            if isinstance(sensor_data, list):
                for data in sensor_data:
                    device_id = data["end_device_ids"]["device_id"]
                    received_at = data.get("received_at")
                    decoded_payload = data["uplink_message"].get("decoded_payload")

                    if decoded_payload:
                        Battery = decoded_payload.get('Battery')
                        EC_S1 = decoded_payload.get('EC_S1')
                        EC_S2 = decoded_payload.get('EC_S2')
                        Moisture_S1 = decoded_payload.get('Moisture_S1')
                        Moisture_S2 = decoded_payload.get('Moisture_S2')
                        Temperature_S1 = decoded_payload.get('Temperature_S1')
                        Temperature_S2 = decoded_payload.get('Temperature_S2')
                        Raw_S1 = decoded_payload.get('Raw_S1')
                        Raw_S2 = decoded_payload.get('Raw_S2')
                        Timestamp = received_at

                        if all(value is not None for value in [Battery, EC_S1, EC_S2, Moisture_S1, Moisture_S2, Temperature_S1, Temperature_S2,Timestamp,Raw_S1,Raw_S2]):
                            try:
                                sensor_property = devices.objects.get(device_id=device_id)
                                create_data = SoilSensor.objects.create(
                                    Battery=float(Battery),
                                    EC_S1=float(EC_S1),
                                    EC_S2=float(EC_S2),
                                    Moisture_S1=float(Moisture_S1),
                                    Moisture_S2=float(Moisture_S2),
                                    Temperature_S1=float(Temperature_S1),
                                    Temperature_S2=float(Temperature_S2),
                                    Raw_S1 =float(Raw_S1),
                                    Raw_S2 =float(Raw_S2),
                                    Timestamp=received_at,  # adjust the field name to lowercase
                                    devices=sensor_property
                                )
                            except devices.DoesNotExist:
                                continue  # Skip if device_id not found in Sensorproperty
                    else:
                        continue  # Skip if decoded_payload is not present
            return Response({'message': 'Data saved successfully.'}, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class WeatherInsertCSV(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = WeatherUploadCSVSerializer

    def post(self, request, format=None):
        csv_file = request.FILES.get('csv_file')
        if 'csv_file' not in request.FILES:
            return Response({"status": "error", "message": "No file uploaded."}, status=400)

        csv_file = request.FILES["csv_file"]

        if csv_file.size == 0 or not csv_file.name.endswith(".csv"):
            return Response({"status": "error", "message": "Only .csv file is supported."}, status=400)

        data_list = []

        try:

            csv_string = csv_file.read().decode('utf-8')
            csv_io = StringIO(csv_string)
            csv_data = csv.reader(csv_io)
            next(csv_data)  # Skip header row

            for row in csv_data:
                if not any(row):
                    break

                timestamp = datetime.strptime(row[0] , "%Y/%m/%d %H:%M:%S %z").strftime('%Y-%m-%d %H:%M:%S')
                wind_dir = row[1]
                solar_rad = row[2]
                leaf_up = row[3]
                leaf_down = row[4]
                battery_voltage = row[5]
                voltage_power_supply = row[6]
                wind_speed = row[7]
                wind_gust = row[8]
                max_rain = row[9]
                daily_rain = row[10]
                rain_counter = row[11]
                temperature = row[12]
                relative_humidity = row[13]
                dew_point = row[14]
                soil_temp = row[15]
                water_content = row[16]

                data = WeatherSensor(
                    Timestamp=timestamp,
                    Wind_dir=wind_dir,
                    Solar_rad=solar_rad,
                    Leaf_up=leaf_up,
                    Leaf_down=leaf_down,
                    Battery_voltage=battery_voltage,
                    Voltage_power_supply=voltage_power_supply,
                    Wind_speed=wind_speed,
                    Wind_gust=wind_gust,
                    Max_rain=max_rain,
                    Daily_rain=daily_rain,
                    Rain_counter=rain_counter,
                    Temperature=temperature,
                    Relative_Humidity=relative_humidity,
                    Dew_point=dew_point,
                    Soil_temp=soil_temp ,
                    water_content = water_content
                )

                data_list.append(data)

            WeatherSensor.objects.bulk_create(data_list)
            return Response({"status": "Success", "message": "Successfully uploaded."})

        except Exception as e:
            return Response({"status": "error", "message": str(e)}, status=400)


class SoilSensorLatestdata(APIView):
    
    def get(self, request, sensor_type=None, start_date=None, end_date=None):
        print(start_date , end_date)
        queryset = SoilSensor.objects.filter(devices__sensor_type=sensor_type, created_at__date__range=(start_date, end_date)).order_by('-Timestamp')
    

        # Convert the location field to GeoJSON
        queryset = queryset.annotate(location_json=AsGeoJSON('devices__location'))

        # Create the GeoJSON data structure
        data = {
            'type': 'FeatureCollection',
            'features': []
        }

        # Convert each SoilSensor object into a GeoJSON feature
        for sensor in queryset:
            original_timestamp = sensor.Timestamp # Assuming the timestamp is already in the correct timezone offset

            # local_timezone = pytz.timezone('Asia/Kolkata')  # Set the local timezone of the application
            original_timezone = pytz.timezone('Asia/Kolkata') 
            localized_timestamp = original_timestamp.astimezone(original_timezone)# Convert to the local timezone




            formatted_timestamp = localized_timestamp.strftime('%Y-%m-%d %I:%M:%S %p')
            feature = {
                'type': 'Feature',
                'geometry': {
                    'type': 'Point',
                    'coordinates': sensor.devices.location.coords,
                },
                'properties': {
                    'Battery': sensor.Battery,
                    'EC_S1': sensor.EC_S1,
                    'EC_S2': sensor.EC_S2,
                    'Moisture_S1': sensor.Moisture_S1,
                    'Moisture_S2': sensor.Moisture_S2,
                    'Temperature_S1': sensor.Temperature_S1,
                    'Temperature_S2': sensor.Temperature_S2,
                    'created_at': sensor.created_at,
                    'device_id': sensor.devices.device_id,
                    'Raw_S1': sensor.Raw_S1,
                    'Raw_S2': sensor.Raw_S2,
                    'Timestamp': formatted_timestamp
                }
            }
            data['features'].append(feature)

        return Response(data)


class SoilSensorLatestDataList(APIView):
    def get(self, request, device_id=None,start_date=None,end_date=None):
        queryset = SoilSensor.objects.filter(devices__device_id=device_id,Timestamp__date__range=(start_date, end_date)).order_by('-Timestamp')

        # Create the GeoJSON data structure
        data = {
            'type': 'FeatureCollection',
            'features': []
        }

        # Convert each SoilSensor object into a GeoJSON feature
        for sensor in queryset:
            location = sensor.devices.location
            coordinates = [location.x, location.y]
            properties = SensorlistSerializer(sensor).data

            # Adjust the timestamp format
            timestamp = properties['Timestamp']
            # formatted_timestamp = datetime.datetime.strptime(timestamp, '%Y-%m-%dT%H:%M:%S.%fZ')
            
         
            # localized_timestamp = formatted_timestamp.replace(tzinfo=datetime.timezone.utc).astimezone(pytz.timezone('UTC'))
            Date = properties['Timestamp'].split("T")
            time = Date[1].split(".")[0]
            properties['Timestamp'] = (Date[0] + " " + time  )
        

            feature = {
                'type': 'Feature',
                'geometry': {
                    'type': 'Point',
                    'coordinates': coordinates
                },
                'properties': properties
            }
            data['features'].append(feature)

        return Response(data)


class SensorGraphAPIView(APIView):
    def get(self, request, device_id, parameter, start_date, end_date):
    
        
        sensor_data = SoilSensor.objects.filter(devices__device_id=device_id,
                                                Timestamp__date__range=[start_date, end_date]).order_by('-Timestamp')
        if sensor_data:
            serializer = SoilSensorGraphSerializer(sensor_data, many=True)
            data = serializer.data


            response_data = {
                'device_id': device_id,
                'parameter': parameter,
                'parameter_data': [],
            }
            # print(response_data)

            for item in data:
                if parameter in item:
                     
                     if parameter in item:
                         Date = item['Timestamp'].split("T")
                         time = Date[1].split(".")[0]
                         response_data['parameter_data'].append({
                            'value': item[parameter],
                            'timestamp': (Date[0] + " " + time  )
                        })
                    
           
            return Response(response_data)
        else:
            return Response({'error': 'No data found for the given parameters.'}, status=404)
    






class DownloadExcelAllSoilSensor(generics.GenericAPIView):
    serializer_class = GeosensorExcelListSerialzier
    parser_classes = [MultiPartParser]
    queryset = devices.objects.all()

    def get(self, request, sensor_type=None, start_date=None, end_date=None):
        try:
            start_date = pd.to_datetime(start_date)
            end_date = pd.to_datetime(end_date)
        
        except Exception as e:
            return Response({'status': 'error', 'message': 'Invalid date format.'}, status=400)
        
        if start_date > end_date:
            return Response({'status': 'error', 'message': 'Invalid date range.'}, status=400)
        
        # Filter the devices by sensor_type
        device_data = self.get_queryset().filter(sensor_type=sensor_type)
        
        sensors = []
        for device in device_data:
            queryset = SoilSensor.objects.filter(devices=device, Timestamp__date__range=(start_date, end_date)).order_by('-Timestamp')
            sensors.extend(queryset)
        
        # Check if any Sensor objects were found
        if not sensors:
            return Response({"message": "No Sensor data found for the specified type and date.", "status": "error"}, status=400)

        # Serialize the data in GeoJSON format
        data = {
            'type': 'FeatureCollection',
            'features': []
        }
        for sensor in sensors:
            location = sensor.devices.location
            feature = {
                'type': 'Feature',
                'geometry': {
                    'type': 'Point',
                    'coordinates': location
                },
                'properties': SensorlistSerializer(sensor).data
            }
            data['features'].append(feature)

        # Convert GeoJSON data to pandas DataFrame
        df = pd.json_normalize(data['features']).drop(['geometry.type'], axis=1)
        df['geometry'] = df['geometry.coordinates'].apply(lambda coords: Point(coords[0], coords[1]))
        df.drop(['geometry.coordinates'], axis=1, inplace=True)
        # df.insert(0, 'Sr. No.', range(1, len(df) + 1))


        

        # Rename columns
        df.rename(columns={
            "properties.Battery": "Battery",
            "properties.EC_S2": "EC_S2",
            "properties.EC_S1": "EC_S1",
            "properties.Temperature_S1": "Temperature_S1",
            "properties.Temperature_S2": "Temperature_S2",
            "properties.Moisture_S1": "Moisture_S1",
            "properties.Moisture_S2": "Moisture_S2",
            "properties.Raw_S1":"Raw_S1",
            "properties.Raw_S2":"Raw_S2",
            "properties.device_id": "device_id",
            "properties.Timestamp": "Timestamp",
        }, inplace=True)

        # Adjust the timestamp format
        df['Timestamp'] = pd.to_datetime(df['Timestamp'])  # Convert to pandas datetime format
        local_timezone = pytz.timezone('Asia/Kolkata')  # Set the local timezone
        df['Timestamp'] = df['Timestamp'].dt.tz_convert(local_timezone)  # Convert to the local timezone
        df['Timestamp'] = df['Timestamp'].dt.strftime('%Y-%m-%d %I:%M:%S %p')  # Format the timestamp

        # Write DataFrame to Excel file
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        df.drop(['type'], axis=1).to_excel(writer, sheet_name='Sheet1', index=False)  # Change index=True to index=False
        worksheet = writer.sheets['Sheet1']
        # worksheet.set_column('A:A', 10)  # Set the width of the Sr. No. column
        writer.book.close()

        # Set the response content type and headers
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Soil_sensor_data_{start_date.date()}_{end_date.date()}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Write the Excel file to the response
        output.seek(0)
        response.write(output.getvalue())

        return response






class DownloadExcelBySoilSensorDeviceid(generics.GenericAPIView):
    serializer_class = GeosensorExcelListSerialzier
    parser_classes = [MultiPartParser]
    queryset = devices.objects.all()

    def get(self, request, sensor_type=None, start_date=None, end_date=None, device_id=None):
        try:
            start_date = pd.to_datetime(start_date)
            end_date = pd.to_datetime(end_date)
        except Exception as e:
            return Response({'status': 'error', 'message': 'Invalid date format.'}, status=400)
        
        if start_date > end_date:
            return Response({'status': 'error', 'message': 'Invalid date range.'}, status=400)
        
        # Filter the devices by sensor_type and device_id
        device_data = self.get_queryset().filter(sensor_type=sensor_type)
        if device_id:
            device_data = device_data.filter(device_id=device_id)
        
        sensors = []
        for device in device_data:
            queryset = SoilSensor.objects.filter(devices=device, Timestamp__date__range=(start_date, end_date)).order_by('-Timestamp')
            sensors.extend(queryset)
        
        # Check if any Sensor objects were found
        if not sensors:
            return Response({"message": "No Sensor data found for the specified type, date, and device.",
                             "status": "error"}, status=400)

        # Serialize the data in GeoJSON format
        data = {
            'type': 'FeatureCollection',
            'features': []
        }
        for sensor in sensors:
            location = sensor.devices.location
            feature = {
                'type': 'Feature',
                'geometry': {
                    'type': 'Point',
                    'coordinates': location
                },
                'properties': SensorlistSerializer(sensor).data
            }
            data['features'].append(feature)

        # Convert GeoJSON data to pandas DataFrame
        df = pd.json_normalize(data['features']).drop(['geometry.type'], axis=1)
        df['geometry'] = df['geometry.coordinates'].apply(lambda coords: Point(coords[0], coords[1]))
  
        df.drop(['geometry.coordinates'], axis=1, inplace=True)

        # df.insert(0, 'Sr. No.', range(1, len(df) + 1))

        # Rename columns
        df.rename(columns={
            "properties.Battery": "Battery",
            "properties.EC_S2": "EC_S2",
            "properties.EC_S1": "EC_S1",
            "properties.Temperature_S1": "Temperature_S1",
            "properties.Temperature_S2": "Temperature_S2",
            "properties.Moisture_S1": "Moisture_S1",
            "properties.Moisture_S2": "Moisture_S2",
            "properties.Raw_S1":"Raw_S1",
            "properties.Raw_S2":"Raw_S2",
            "properties.device_id": "device_id",
            "properties.Timestamp": "Timestamp",
        }, inplace=True)

        # Adjust the timestamp format
        df['Timestamp'] = pd.to_datetime(df['Timestamp'])  # Convert to pandas datetime format
        local_timezone = pytz.timezone('Asia/Kolkata')  # Set the local timezone
        df['Timestamp'] = df['Timestamp'].dt.tz_convert(local_timezone)  # Convert to the local timezone
        df['Timestamp'] = df['Timestamp'].dt.strftime('%Y-%m-%d %I:%M:%S %p')  # Format the timestamp

        # Write DataFrame to Excel file
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        df.drop(['type'], axis=1).to_excel(writer, sheet_name='Sheet1', index=False)
        writer.book.close()

        # Set the response content type and headers
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Soil_sensor_data_{start_date.date()}_{end_date.date()}_{device_id}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Write the Excel file to the response
        output.seek(0)
        response.write(output.getvalue())

        return response




